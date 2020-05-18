# Section 1 - Loading our Libraries
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference

df = pd.read_excel('sample_pivot.xlsx', parse_dates=['Date'])
print(df.head())

filtered = df[df['Region'] == 'East']
quarterly_sales = pd.pivot_table(filtered, index = filtered['Date'].dt.quarter, columns = 'Type', values = 'Sales', aggfunc='sum')

print("Quarterly Sales Pivot Table:")
print(quarterly_sales.head())

file_path = 'quarterly_sales.xlsx'
quarterly_sales.to_excel(file_path, sheet_name = 'Quarterly Sales', startrow=3)

wb = load_workbook(file_path)
sheet1 = wb['Quarterly Sales']

# Section 06 - Formatting the First Sheet
sheet1['A1'] = 'Quarterly Sales'
sheet1['A2'] = 'datagy.io'
sheet1['A4'] = 'Quarter'

sheet1['A1'].style = 'Title'
sheet1['A2'].style = 'Headline 2'

for i in range(5, 9):
    sheet1[f'B{i}'].style='Currency'
    sheet1[f'C{i}'].style='Currency'
    sheet1[f'D{i}'].style='Currency'

# Section 07 - Adding a Bar Chart
bar_chart = BarChart()
data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(categories)
sheet1.add_chart(bar_chart, "F4")

bar_chart.title = 'Sales by Type'
bar_chart.style = 3
wb.save(filename = file_path)


regions = list(df['Region'].unique())
folder_path = ""

for region in regions:
    filtered = df[df['Region'] == f'{region}']
    quarterly_sales = pd.pivot_table(filtered, index = filtered['Date'].dt.quarter, columns = 'Type', values = 'Sales', aggfunc='sum')
    file_path = f"{region}.xlsx"
    quarterly_sales.to_excel(file_path, sheet_name = 'Quarterly Sales', startrow=3)
    
    wb = load_workbook(file_path)
    sheet1 = wb['Quarterly Sales']
    
    sheet1['A1'] = 'Quarterly Sales'
    sheet1['A2'] = 'datagy.io'
    sheet1['A4'] = 'Quarter'

    sheet1['A1'].style = 'Title'
    sheet1['A2'].style = 'Headline 2'

    for i in range(5, 10):
        sheet1[f'B{i}'].style='Currency'
        sheet1[f'C{i}'].style='Currency'
        sheet1[f'D{i}'].style='Currency'

    bar_chart = BarChart()
    data = Reference(sheet1, min_col=2, max_col=4, min_row=4, max_row=8)
    categories = Reference(sheet1, min_col=1, max_col=1, min_row=5, max_row=8)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    sheet1.add_chart(bar_chart, "F4")

    bar_chart.title = 'Sales by Type'
    bar_chart.style = 3
    wb.save(file_path)